import sqlite3
import os
import shutil

DB_NAME = 'database.db'

def create_database():
    """Создаёт все таблицы при первом запуске"""
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        
        # 1. Таблица пользователей
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                login TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL,
                full_name TEXT NOT NULL,
                role TEXT NOT NULL
            )
        ''')
        
        # 2. Таблица поставщиков
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS suppliers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL
            )
        ''')
        
        # 3. Таблица товаров (с артикулом)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                article TEXT UNIQUE,
                name TEXT NOT NULL,
                category TEXT NOT NULL,
                description TEXT,
                manufacturer TEXT,
                supplier_id INTEGER,
                price REAL NOT NULL CHECK(price >= 0),
                unit TEXT,
                quantity INTEGER NOT NULL DEFAULT 0 CHECK(quantity >= 0),
                discount REAL DEFAULT 0,
                image_path TEXT,
                FOREIGN KEY (supplier_id) REFERENCES suppliers(id)
            )
        ''')
        
        # 4. Таблица заказов
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                article TEXT NOT NULL,
                status TEXT NOT NULL,
                pickup_address TEXT NOT NULL,
                order_date TEXT NOT NULL,
                issue_date TEXT
            )
        ''')
        
        # 5. Таблица связи заказов и товаров
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS order_products (
                order_id INTEGER,
                product_id INTEGER,
                quantity INTEGER,
                FOREIGN KEY (order_id) REFERENCES orders(id) ON DELETE CASCADE,
                FOREIGN KEY (product_id) REFERENCES products(id)
            )
        ''')
        
        # 6. Добавляем тестовых пользователей
        users = [
            ('admin', 'admin123', 'Администратор Системы', 'admin'),
            ('manager', 'manager123', 'Петров Иван Сергеевич', 'manager'),
            ('client', 'client123', 'Сидорова Анна Ивановна', 'client'),
        ]
        
        for user in users:
            cursor.execute('SELECT COUNT(*) FROM users WHERE login = ?', (user[0],))
            if cursor.fetchone()[0] == 0:
                cursor.execute('''
                    INSERT INTO users (login, password, full_name, role)
                    VALUES (?, ?, ?, ?)
                ''', user)
        
        conn.commit()
    
    os.makedirs('images', exist_ok=True)
    print("База данных создана успешно!")


def import_products_from_excel(file_path):
    """Импортирует товары из Excel файла"""
    try:
        import openpyxl
    except ImportError:
        print("Установите openpyxl: pip install openpyxl")
        return 0
    
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    added = 0
    excel_dir = os.path.dirname(file_path)
    images_source_dir = 'images_source'
    os.makedirs('images', exist_ok=True)
    
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[1]:
                continue
            
            # Читаем артикул из колонки A (индекс 0)
            article = str(row[0]).strip() if row[0] else ""
            name = str(row[1]).strip()
            unit = str(row[2]).strip() if row[2] else "шт"
            
            try: 
                price = float(row[3]) if row[3] else 0
            except: 
                price = 0
            
            supplier_name = str(row[4]).strip() if row[4] else ""
            manufacturer = str(row[5]).strip() if row[5] else ""
            category = str(row[6]).strip() if row[6] else ""
            
            try: 
                discount = float(row[7]) if row[7] else 0
            except: 
                discount = 0
            
            try: 
                qty = int(row[8]) if row[8] else 0
            except: 
                qty = 0
            
            description = str(row[9]).strip() if row[9] else ""
            image_filename = str(row[10]).strip() if row[10] else ""
            
            # Обработка фото
            image_path = ""
            if image_filename:
                source_image = None
                candidate1 = os.path.join(images_source_dir, image_filename)
                candidate2 = os.path.join(excel_dir, image_filename)
                
                if os.path.exists(candidate1):
                    source_image = candidate1
                elif os.path.exists(candidate2):
                    source_image = candidate2
                
                if source_image:
                    dest_path = os.path.join('images', image_filename)
                    shutil.copy2(source_image, dest_path)
                    image_path = dest_path
                    print(f"  Фото скопировано: {image_filename}")
                else:
                    print(f"  Фото не найдено: {image_filename}")
            
            if not name:
                continue
            
            # Получаем или создаём поставщика
            supplier_id = None
            if supplier_name:
                cursor.execute('SELECT id FROM suppliers WHERE name = ?', (supplier_name,))
                sup = cursor.fetchone()
                if sup:
                    supplier_id = sup[0]
                else:
                    cursor.execute('INSERT INTO suppliers (name) VALUES (?)', (supplier_name,))
                    supplier_id = cursor.lastrowid
            
            # Проверяем дубликаты по артикулу
            if article:
                cursor.execute('SELECT id FROM products WHERE article = ?', (article,))
                if cursor.fetchone():
                    print(f"Товар с артикулом {article} уже существует, пропускаем")
                    continue
            
            cursor.execute('''
                INSERT INTO products (name, category, description, manufacturer, supplier_id,
                                      price, unit, quantity, discount, image_path, article)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (name, category, description, manufacturer, supplier_id,
                  price, unit, qty, discount, image_path, article))
            added += 1
            print(f"Добавлен товар: {name} (артикул: {article})")
        
        conn.commit()
    
    print(f"\nИмпортировано товаров: {added}")
    return added


def import_orders_from_excel(file_path):
    """Импортирует заказы из Excel файла с подстановкой адресов"""
    try:
        import openpyxl
    except ImportError:
        print("Установите openpyxl: pip install openpyxl")
        return 0
    
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
    except Exception as e:
        print(f"Ошибка чтения файла: {e}")
        return 0
    
    # Загружаем адреса из файла пункты_выдачи.xlsx
    pickup_addresses = {}
    try:
        addr_workbook = openpyxl.load_workbook('data/пункты_выдачи.xlsx')
        addr_sheet = addr_workbook.active
        # Пронумеруем адреса с 1
        for idx, row in enumerate(addr_sheet.iter_rows(min_row=1, values_only=True), start=1):
            if row and row[0]:
                pickup_addresses[str(idx)] = str(row[0])
        print(f"Загружено адресов: {len(pickup_addresses)}")
    except Exception as e:
        print(f"Не удалось загрузить адреса: {e}")
    
    added_orders = 0
    added_products = 0
    skipped_orders = 0
    
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            
            # Читаем данные заказа
            order_num = str(row[0]).strip() if row[0] else ""
            products_str = str(row[1]) if row[1] else ""
            order_date = row[2] if row[2] else ""
            delivery_date = row[3] if row[3] else ""
            address_num = str(row[4]).strip() if row[4] else ""
            client_name = row[5] if row[5] else ""
            pickup_code = row[6] if row[6] else ""
            status = str(row[7]) if row[7] else "Новый"
            
            if not order_num:
                continue
            
            # Подставляем полный адрес по номеру
            pickup_address = pickup_addresses.get(address_num, address_num)
            
            # Проверяем, есть ли уже такой заказ
            cursor.execute('SELECT id FROM orders WHERE article = ?', (order_num,))
            if cursor.fetchone():
                print(f"Заказ {order_num} уже существует, пропускаем")
                skipped_orders += 1
                continue
            
            # Добавляем заказ
            cursor.execute('''
                INSERT INTO orders (article, status, pickup_address, order_date, issue_date)
                VALUES (?, ?, ?, ?, ?)
            ''', (order_num, status, pickup_address, str(order_date), str(delivery_date) if delivery_date else None))
            
            order_id = cursor.lastrowid
            
            # Разбираем товары
            if products_str:
                parts = [p.strip() for p in products_str.split(',')]
                
                for i in range(0, len(parts), 2):
                    if i + 1 < len(parts):
                        product_article = parts[i]
                        try:
                            quantity = int(parts[i + 1])
                        except ValueError:
                            quantity = 1
                        
                        cursor.execute('SELECT id FROM products WHERE article = ?', (product_article,))
                        product = cursor.fetchone()
                        
                        if product:
                            product_id = product[0]
                            cursor.execute('''
                                INSERT INTO order_products (order_id, product_id, quantity)
                                VALUES (?, ?, ?)
                            ''', (order_id, product_id, quantity))
                            added_products += 1
                            print(f"  Товар {product_article} x{quantity} добавлен в заказ {order_num}")
                        else:
                            print(f"  Товар с артикулом {product_article} не найден, пропускаем")
            
            added_orders += 1
            print(f"Добавлен заказ {order_num} (адрес: {pickup_address})")
        
        conn.commit()
    
    print(f"\nИмпортировано заказов: {added_orders}, пропущено: {skipped_orders}")
    print(f"Импортировано позиций: {added_products}")
    return added_orders

def import_users_from_excel(file_path):
    """Импортирует пользователей из Excel файла"""
    try:
        import openpyxl
    except ImportError:
        print("Установите openpyxl: pip install openpyxl")
        return 0
    
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
    except Exception as e:
        print(f"Ошибка чтения файла: {e}")
        return 0
    
    added = 0
    skipped = 0
    
    # Определяем заголовки
    headers = {}
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(1, col).value
        if header:
            headers[header] = col
    
    print(f"Найденные колонки: {list(headers.keys())}")
    
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            
            role_ru = str(row[headers.get("Роль сотрудника", 0) - 1]) if "Роль сотрудника" in headers else ""
            full_name = str(row[headers.get("ФИО", 1) - 1]) if "ФИО" in headers else ""
            login = str(row[headers.get("Логин", 2) - 1]) if "Логин" in headers else ""
            password = str(row[headers.get("Пароль", 3) - 1]) if "Пароль" in headers else ""
            
            if not login or not full_name:
                continue
            
            role_map = {
                "Администратор": "admin",
                "Менеджер": "manager",
                "Авторизированный клиент": "client",
                "Клиент": "client"
            }
            role = role_map.get(role_ru, "client")
            
            cursor.execute('SELECT id FROM users WHERE login = ?', (login,))
            if cursor.fetchone():
                print(f"Пользователь {login} уже существует, пропускаем")
                skipped += 1
                continue
            
            cursor.execute('''
                INSERT INTO users (login, password, full_name, role)
                VALUES (?, ?, ?, ?)
            ''', (login, password, full_name, role))
            added += 1
            print(f"Добавлен пользователь: {login} ({full_name}) - {role}")
        
        conn.commit()
    
    print(f"\nИмпортировано пользователей: {added}, пропущено: {skipped}")
    return added